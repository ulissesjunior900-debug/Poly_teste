import os
import re
import unicodedata
from io import BytesIO
from decimal import Decimal
from datetime import datetime
from flask import current_app
from openpyxl import load_workbook
import win32com.client as win32
from werkzeug.utils import secure_filename
import tempfile




def formatar_valor(valor, casas=2):
    return f"{Decimal(valor).quantize(Decimal('1.' + '0' * casas), rounding=ROUND_HALF_UP)}"

def formatar_data(data_iso):
    if '-' in data_iso:
        partes = data_iso.split('-')
        return f"{partes[2]}/{partes[1]}/{partes[0]}"
    return data_iso

# Mapa de meses PT-BR abreviado e por extenso:
# Mapa de meses PT-BR abreviado e por extenso:
mes_abrev_pt = {
    1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR', 5: 'MAI', 6: 'JUN',
    7: 'JUL', 8: 'AGO', 9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
}

mes_extenso_pt = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho',
    7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}

def preencher_sp_dinamicamente(sp_obj, valor_eur, cotacao, vencimento, retencao=0, valores_adicionais=None, mes=None, ano=None, artista=None):
    try:
        import unicodedata
        import re
        from decimal import Decimal
        from datetime import datetime
        import os
        import traceback
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
        from io import BytesIO

        # Validação inicial dos parâmetros obrigatórios
        if not all([sp_obj, valor_eur is not None, cotacao is not None, vencimento]):
            raise ValueError("Parâmetros obrigatórios não fornecidos")

        # Inicializa valores_adicionais se None
        if valores_adicionais is None:
            valores_adicionais = {}

        # ==== CAPTURA DOS DADOS ====
        artista_nome = artista or valores_adicionais.get('artista', '').strip()
        mes_calculo = str(mes or valores_adicionais.get('mes', '')).strip()
        ano_calculo = str(ano or valores_adicionais.get('ano', '')).strip()

        def extrair_mes_ano(texto):
            if not texto:
                return None, None
            padrao = r'(?:jan|janeiro|fev|fevereiro|mar|março|abr|abril|mai|maio|jun|junho|jul|julho|ago|agosto|set|setembro|out|outubro|nov|novembro|dez|dezembro)[/ ]*(\d{4})'
            match = re.search(padrao, texto.lower())
            if match:
                mes = match.group(0).split('/')[0].strip()
                ano = match.group(1)
                return mes, ano
            return None, None

        # Fallback: pegar mes/ano se não veio no texto
        if not mes_calculo:
            mes_calculo = valores_adicionais.get('mes', '').strip()
        if not ano_calculo:
            ano_calculo = valores_adicionais.get('ano', '').strip()

        # Fallback final: se veio calculos_ids → buscar no banco
        calculos_ids_raw = valores_adicionais.get('calculos_ids')
        if calculos_ids_raw:
            from models import CalculoSalvo, CalculoEspecialSalvo, CalculoAssisaoSalvo, db
            
            ids = []
            if isinstance(calculos_ids_raw, str):
                ids = [int(i) for i in calculos_ids_raw.split(',') if i.strip().isdigit()]
            elif isinstance(calculos_ids_raw, list):
                ids = [int(i) for i in calculos_ids_raw if str(i).isdigit()]
            
            if ids:
                calc_id = ids[-1]  # pega o último cálculo para preencher mes/ano/artista
                calc_obj = None
                for model in [CalculoSalvo, CalculoEspecialSalvo, CalculoAssisaoSalvo]:
                    calc_obj = db.session.query(model).filter_by(id=calc_id).first()
                    if calc_obj:
                        break
                
                if calc_obj:
                    if not artista_nome:
                        if hasattr(calc_obj, 'artista') and isinstance(calc_obj.artista, str):
                            artista_nome = calc_obj.artista
                        elif hasattr(calc_obj, 'artista_especial') and calc_obj.artista_especial:
                            artista_nome = calc_obj.artista_especial.nome
                        else:
                            artista_nome = 'Artista não informado'
                    if not mes_calculo:
                        mes_calculo = str(calc_obj.mes)
                    if not ano_calculo:
                        ano_calculo = str(calc_obj.ano)

        # Validação final
        if not all([mes_calculo, ano_calculo, artista_nome]):
            raise ValueError("Dados incompletos: mês, ano ou artista não fornecidos")

        # ==== CONVERSÃO DE MÊS ====
        mes_nome_curto_map = {
            1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR', 5: 'MAI', 6: 'JUN',
            7: 'JUL', 8: 'AGO', 9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
        }

        mes_nome_pt_map = {
            1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho',
            7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
        }

        mes_map = {
            'jan': 1, 'janeiro': 1,
            'fev': 2, 'fevereiro': 2, 'fevereiro': 2, 'fevereiro': 2, 'fevr': 2,
            'mar': 3, 'marco': 3, 'março': 3,
            'abr': 4, 'abril': 4,
            'mai': 5, 'maio': 5,
            'jun': 6, 'junho': 6,
            'jul': 7, 'julho': 7,
            'ago': 8, 'agosto': 8,
            'set': 9, 'setembro': 9,
            'out': 10, 'outubro': 10,
            'nov': 11, 'novembro': 11,
            'dez': 12, 'dezembro': 12
        }

        try:
            mes_num = int(mes_calculo)
        except (ValueError, TypeError):
            mes_normalizado = unicodedata.normalize('NFKD', str(mes_calculo).lower()).encode('ASCII', 'ignore').decode()
            mes_normalizado = re.sub(r'[^a-z]', '', mes_normalizado)
            mes_num = mes_map.get(mes_normalizado)
            if mes_num is None:
                raise ValueError(f"Mês inválido: {mes_calculo}")

        cabecalho_mes_ano = f"{mes_nome_curto_map.get(mes_num)} / {ano_calculo}"
        periodo_completo = f"{mes_nome_pt_map.get(mes_num)} / {ano_calculo}"

        # ==== COTAÇÃO EM +2 MESES ====
        mes_cotacao = mes_num + 2
        ano_cotacao = int(ano_calculo)
        if mes_cotacao > 12:
            mes_cotacao -= 12
            ano_cotacao += 1

        texto_cotacao = f"Cotação em 11/{mes_cotacao:02d}/{ano_cotacao}"

        # ==== CÁLCULO DE VALORES ====
        valor_eur_decimal = Decimal(str(valor_eur))
        cotacao_decimal = Decimal(str(cotacao))
        valor_brl = valor_eur_decimal * cotacao_decimal

        # Conversão segura de retencao
        try:
            retencao_decimal = Decimal(str(retencao).replace(',', '.'))
        except Exception:
            retencao_decimal = Decimal('0')

        # Aplica a retenção apenas se for maior que zero
        if retencao_decimal > Decimal('0'):
            valor_retencao = valor_brl * retencao_decimal / Decimal('100')
        else:
            valor_retencao = Decimal('0')

        valor_liquido = valor_brl - valor_retencao

        tem_herdeiro = valores_adicionais and valores_adicionais.get("herdeiros")
        valor_herdeiro = None
        texto_herdeiro = None

        if tem_herdeiro:
            herdeiro_info = list(valores_adicionais["herdeiros"].items())[0]
            texto_herdeiro = f"{herdeiro_info[0]} {herdeiro_info[1]}%"
            percentual_herdeiro = Decimal(str(herdeiro_info[1])) / 100
            valor_herdeiro = valor_liquido * percentual_herdeiro
            valor_liquido = valor_herdeiro

        # ==== CARREGAR SP ====
        caminho_template = os.path.join(current_app.root_path, sp_obj.caminho.replace('/', os.sep))
        if not os.path.isfile(caminho_template):
            raise FileNotFoundError(f"Template não encontrado em: {caminho_template}")

        wb = load_workbook(caminho_template)
        ws_sp = wb["SP"] if "SP" in wb.sheetnames else wb.active
        ws_resumo = wb["Resumo"] if "Resumo" in wb.sheetnames else None

        def get_top_left_cell(cell, ws):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return ws.cell(merged_range.min_row, merged_range.min_col)
            return cell

        for row in ws_sp.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    texto_original = str(cell.value)
                    texto_normalizado = unicodedata.normalize('NFKD', texto_original.lower()).encode('ASCII', 'ignore').decode()
                    texto_normalizado = re.sub(r'[\n\r\t]', ' ', texto_normalizado)
                    texto_normalizado = re.sub(r'\s+', ' ', texto_normalizado).strip()

                    top_left_cell = get_top_left_cell(cell, ws_sp)

                    if 'vencimento' in texto_normalizado:
                        try:
                            data_venc_formatada = datetime.strptime(vencimento, "%Y-%m-%d").strftime("%d/%m/%Y")
                            ws_sp.cell(cell.row, cell.column + 1).value = data_venc_formatada
                        except:
                            ws_sp.cell(cell.row, cell.column + 1).value = "01/01/2025"

                    elif 'valor a pagar' in texto_normalizado:
                        ws_sp.cell(cell.row, cell.column + 1).value = float(valor_liquido)

                    elif re.search(r'(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)\s*[/-]?\s*\d{4}', texto_original, re.IGNORECASE):
                        top_left_cell.value = cabecalho_mes_ano

                    elif 'valor unit' in texto_normalizado:
                        ws_sp.cell(cell.row + 1, cell.column).value = float(valor_eur_decimal)

                    elif re.search(r'cotacao', texto_normalizado, re.IGNORECASE):
                        top_left_cell.value = texto_cotacao
                        for offset in [1, -1]:
                            adj_cell = ws_sp.cell(top_left_cell.row, top_left_cell.column + offset)
                            if adj_cell.value is None or isinstance(adj_cell.value, (int, float)):
                                adj_cell.value = float(cotacao_decimal)
                                break

                    elif 'total' in texto_normalizado and not 'valor total' in texto_normalizado:
                        ws_sp.cell(cell.row, cell.column + 1).value = float(valor_brl)

                    elif 'retenção iss' in texto_normalizado.lower() and retencao_decimal > Decimal('0'):
                        top_left_cell.value = f"RETENÇÃO ISS {retencao}%"
                        for offset in [1, -1]:
                            adj_cell = ws_sp.cell(top_left_cell.row, top_left_cell.column + offset)
                            if adj_cell.value is None or isinstance(adj_cell.value, (int, float)):
                                adj_cell.value = float(-valor_retencao)
                                break

                    elif tem_herdeiro and ('herdeiro' in texto_normalizado or 'filho' in texto_normalizado or 'esposa' in texto_normalizado or 'filha' in texto_normalizado):
                        top_left_cell.value = texto_herdeiro
                        for offset in [1, -1]:
                            adj_cell = ws_sp.cell(top_left_cell.row, top_left_cell.column + offset)
                            if adj_cell.value is None or isinstance(adj_cell.value, (int, float)):
                                adj_cell.value = float(valor_herdeiro)
                                break

                    elif 'valor total' in texto_normalizado:
                        ws_sp.cell(cell.row, cell.column + 1).value = float(valor_liquido)

        if ws_resumo:
            ws_resumo["A5"].value = mes_nome_pt_map.get(datetime.now().month)
            ws_resumo["A8"].value = artista_nome
            ws_resumo["A10"].value = periodo_completo
            ws_resumo["B10"].value = float(valor_eur_decimal)
            ws_resumo["A12"].value = texto_cotacao
            ws_resumo["B12"].value = float(cotacao_decimal)
            ws_resumo["B13"].value = float(valor_brl)

            if retencao_decimal > Decimal('0'):
                ws_resumo["A15"].value = f"RETENÇÃO ISS {retencao_decimal}%"
                ws_resumo["B15"].value = float(-valor_retencao)
            elif tem_herdeiro:
                ws_resumo["A15"].value = texto_herdeiro
                ws_resumo["B15"].value = float(valor_herdeiro)

            if retencao_decimal > Decimal('0') or tem_herdeiro:
                ws_resumo["A16"].value = "Valor Líquido"
                ws_resumo["B16"].value = float(valor_liquido)

        # ==== INSERÇÃO DE LOGO E ASSINATURA (CORREÇÃO DO ERRO) ====
        try:
            def cm_to_px(cm):
                return int(cm * 37.8)

            imagens_dir = os.path.join(current_app.root_path, 'static', 'imagens')

            # Carregar imagens em bytes
            logo_bytes = None
            assinatura_bytes = None

            caminho_logo = os.path.join(imagens_dir, 'logo.png')
            if os.path.isfile(caminho_logo):
                with open(caminho_logo, 'rb') as f:
                    logo_bytes = f.read()

            caminho_assinatura = os.path.join(imagens_dir, 'assinatura.png')
            if os.path.isfile(caminho_assinatura):
                with open(caminho_assinatura, 'rb') as f:
                    assinatura_bytes = f.read()

            # Lista para manter os buffers ativos
            imagem_buffers = []

            # Inserir logo
            if logo_bytes:
                logo_buffer = BytesIO(logo_bytes)
                imagem_buffers.append(logo_buffer)  # Manter referência
                logo_img = Image(logo_buffer)
                logo_img.width = cm_to_px(10.50)
                logo_img.height = cm_to_px(4.26)
                ws_sp.add_image(logo_img, 'B2')

            # Inserir assinatura
            if assinatura_bytes:
                assinatura_buffer = BytesIO(assinatura_bytes)
                imagem_buffers.append(assinatura_buffer)  # Manter referência
                assinatura_img = Image(assinatura_buffer)
                assinatura_img.width = cm_to_px(2.27)
                assinatura_img.height = cm_to_px(1.38)

                solicitante_cell = None
                for row in ws_sp.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            texto = unicodedata.normalize('NFKD', cell.value.lower()).encode('ASCII', 'ignore').decode()
                            texto = re.sub(r'[\n\r\t]', ' ', texto)
                            texto = re.sub(r'\s+', ' ', texto).strip()

                            if 'solicitante.' in texto:
                                solicitante_cell = cell
                                break
                    if solicitante_cell:
                        break

                if solicitante_cell:
                    col_letter = solicitante_cell.column_letter
                    row_number = solicitante_cell.row + 1
                    posicao_assinatura = f"{col_letter}{row_number}"
                    ws_sp.add_image(assinatura_img, posicao_assinatura)

        except Exception as img_error:
            print(f"[ERROR] Erro ao inserir logo ou assinatura: {str(img_error)}")
        
        # GERAR ARQUIVO EM MEMÓRIA (PRINCIPAL)
        memoria = BytesIO()
        wb.save(memoria)
        memoria.seek(0)

        # Salvar cópia no disco apenas se necessário para PDF
        caminho_saida = None
        output_dir = os.path.join(current_app.root_path, "static", "downloads")
        os.makedirs(output_dir, exist_ok=True)
        nome_arquivo_saida = f"SP_{sp_obj.nome_arquivo.replace('.xlsx', '')}_preenchida.xlsx"
        caminho_saida = os.path.join(output_dir, nome_arquivo_saida)
        
        with open(caminho_saida, 'wb') as f:
            f.write(memoria.getvalue())
        memoria.seek(0)  # Resetar o buffer para uso futuro

        nome_arquivo = os.path.basename(sp_obj.caminho)

        return {
            "arquivo_memoria": memoria,
            "nome_arquivo": nome_arquivo,
            "mes_ano": cabecalho_mes_ano,
            "periodo_completo": periodo_completo,
            "caminho_excel": caminho_saida
        }

    except Exception as e:
        print(f"[ERROR] {str(e)}")
        print(traceback.format_exc())
        raise


def gerar_sp_pdf_com_preenchimento(
    sp_obj,
    calculos,
    cotacao_valor,
    vencimento,
    status_pagamento,
    retencao=None,
    herdeiros_dict=None,
    anexos={}
):
    try:
        from utils_sp import preencher_sp_dinamicamente, gerar_pdf_completo_da_sp
        import tempfile
        import shutil
        import os
        from io import BytesIO
        from werkzeug.utils import secure_filename
        from PyPDF2 import PdfMerger
        from PIL import Image
        from fpdf import FPDF
        import win32com.client

        # Verifica se há pelo menos um cálculo
        if not calculos or not isinstance(calculos, list):
            raise ValueError("Lista de cálculos está vazia ou inválida.")

        # Extrair o último cálculo como base para mês, ano e artista
        calc_base = calculos[-1]
        valor_total_eur = sum(Decimal(str(c['valor_eur'])) for c in calculos)

        valores_adicionais = {
            "artista": calc_base.get("artista"),
            "mes": calc_base.get("mes"),
            "ano": calc_base.get("ano"),
            "herdeiros": herdeiros_dict or {},
            "calculos_ids": [c.get("id") for c in calculos if "id" in c]
        }

        # Preencher planilha Excel em memória
        resultado = preencher_sp_dinamicamente(
            sp_obj=sp_obj,
            valor_eur=valor_total_eur,
            cotacao=cotacao_valor,
            vencimento=vencimento,
            retencao=retencao,
            valores_adicionais=valores_adicionais,
            artista=calc_base.get("artista"),
            mes=calc_base.get("mes"),
            ano=calc_base.get("ano")
        )

        caminho_excel = resultado.get("caminho_excel")
        if not caminho_excel or not os.path.isfile(caminho_excel):
            # Se não foi salvo em disco, tentar usar o buffer para criar um temp
            if resultado.get("arquivo_memoria"):
                memoria = resultado["arquivo_memoria"]
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    tmp.write(memoria.getvalue())
                    caminho_excel = tmp.name
            else:
                raise FileNotFoundError("Arquivo da SP não foi gerado corretamente.")

        # Gerar PDF com base na planilha preenchida
        pdf_final, temp_dir = gerar_pdf_completo_da_sp(caminho_excel, anexos)

        # Ler o PDF gerado para memória
        with open(pdf_final, 'rb') as f:
            pdf_bytes = f.read()
        
        pdf_memoria = BytesIO(pdf_bytes)
        pdf_memoria.seek(0)

        # Limpar temporários
        shutil.rmtree(temp_dir, ignore_errors=True)
        # Se o arquivo Excel foi criado temporariamente, excluímos
        if not resultado.get("caminho_excel"):
            try:
                os.unlink(caminho_excel)
            except:
                pass

        return {
            "pdf_memoria": pdf_memoria,
            "nome_arquivo": sp_obj.nome_arquivo.replace('.xlsx', '.pdf'),
            "mes_ano": resultado.get("mes_ano"),
            "periodo_completo": resultado.get("periodo_completo"),
        }

    except Exception as e:
        print(f"[ERRO] Erro interno ao gerar PDF da SP: {e}")
        raise

def gerar_pdf_completo_da_sp(caminho_excel, anexos={}):
    """
    Gera um PDF final com base nas abas "SP" e "Resumo" da planilha Excel preenchida,
    mais os anexos (como NF e Visto).

    Retorna o caminho completo do PDF gerado.
    """

    def salvar_aba_como_imagem(excel_path, aba_nome, saida_img):
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(os.path.abspath(excel_path))
        try:
            ws = wb.Worksheets(aba_nome)
            ws.Range("A1:Z50").CopyPicture(Format=2)
            chart = excel.Charts.Add()
            chart.Paste()
            chart.Export(os.path.abspath(saida_img))
            chart.Delete()
        finally:
            wb.Close(False)
            excel.Quit()

    def montar_pdf(imagens, caminho_pdf_saida):
        pdf = FPDF()
        for img_path in imagens:
            if img_path.lower().endswith(".pdf"):
                continue

            img = Image.open(img_path)
            w_px, h_px = img.size
            dpi = 96
            px_to_mm = 25.4 / dpi
            w_mm = w_px * px_to_mm
            h_mm = h_px * px_to_mm

            pdf.add_page()
            if "SP" in os.path.basename(img_path):
                x = max((210 - w_mm) / 2, 0)
                y = max((297 - h_mm) / 2, 0)
                pdf.image(img_path, x=x, y=y, w=w_mm, h=h_mm)
            else:
                pdf.image(img_path, x=0, y=0, w=210, h=297)

        pdf.output(caminho_pdf_saida)

    # Criar temp dir
    temp_dir = tempfile.mkdtemp()
    imagens = []

    for aba in ['SP', 'Resumo']:
        img_path = os.path.join(temp_dir, f"{aba}.png")
        salvar_aba_como_imagem(caminho_excel, aba, img_path)
        imagens.append(img_path)

    # Salvar anexos
    for nome, file_storage in anexos.items():
        if file_storage:
            nome_seguro = secure_filename(file_storage.filename)
            caminho_anexo = os.path.join(temp_dir, nome_seguro)
            file_storage.save(caminho_anexo)
            imagens.append(caminho_anexo)

    # PDF base
    pdf_base = os.path.join(temp_dir, "SP_Base.pdf")
    montar_pdf(imagens, pdf_base)

    # PDF final
    pdf_final = os.path.join(temp_dir, "SP_Completa.pdf")
    merger = PdfMerger()
    merger.append(pdf_base)

    for item in imagens:
        if item.lower().endswith(".pdf"):
            merger.append(item)

    merger.write(pdf_final)
    merger.close()

    return pdf_final, temp_dir  # temp_dir para a rota poder limpar depois